import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

def read_excel_range():
    # Load the workbook
    wb = openpyxl.load_workbook('C:\\Users\\consultants\\Silver Crystal Sports\\ABS - Documents\\ABS - Batch Allocate and Printing\\WHS Replenishment Tool v13.xlsm', data_only=True)
    
    # Select the specific sheet
    sheet = wb['Auto Pull List']
    
    # Initialize variables for our range
    start_col = 'N'
    end_col = 'T'
    
    # Convert column letters to numbers (1-based)
    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col)
    
    # Find the last populated row in column N
    max_row = 1
    for row in sheet[start_col]:
        if row.value is not None:
            max_row = row.row
        else:
            break
    
    # Create a list to store our data
    data = []
    
    # Read the data from the sheet
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(start_col_idx, end_col_idx + 1):
            cell_value = sheet[f'{get_column_letter(col)}{row}'].value

            #if cell_value is a string and contains a " remove it
            if isinstance(cell_value, str):
                cell_value = cell_value.replace('"', '')

            row_data.append(cell_value)
        data.append(row_data)
    
    # Get the header row (first row)
    headers = data[0]
    
    # Create DataFrame with the data, excluding the header row
    df = pd.DataFrame(data[1:], columns=headers)
    
    # Close the workbook
    wb.close()
    
    return df

def read_specialPickList():
    wb = openpyxl.load_workbook("C:\\Users\\consultants\\Silver Crystal Sports\\ABS - Documents\\ABS - Batch Allocate and Printing\\WHS Replenishment Tool v13.xlsm", data_only=True)
    sheet = wb['Special Item Pick List']

    # Initialize variables for our range
    start_col = 'G'
    end_col = 'M'
    
    # Convert column letters to numbers (1-based)
    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col)
    
    # Find the last populated row in column N
    max_row = 1
    blankCount = 0
    for row in sheet["G"]:
        if row.value is not None:
            if row.row > max_row:
                max_row = row.row
        else:
            blankCount +=1
            if blankCount >= 10:
                break

    # Create a list to store our data
    data = []
    
    # Read the data from the sheet
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(start_col_idx, end_col_idx + 1):
            cell_value = sheet[f'{get_column_letter(col)}{row}'].value
            row_data.append(cell_value)
        data.append(row_data)
    

    # Get the header row (first row)
    headers = data[1]

    print(headers)
    
    # Create DataFrame with the data, excluding the header row
    df = pd.DataFrame(data[2:], columns=headers)
    
    # Close the workbook
    wb.close()
    
    return df


def read_callUp_list():
    wb = openpyxl.load_workbook("C:\\Users\\consultants\\Silver Crystal Sports\\ABS - Documents\\ABS - Batch Allocate and Printing\\WHS Replenishment Tool v13.xlsm", data_only=True)
    sheet = wb['BWX Box-Skid Call Up']
    
    # Initialize variables for our range
    start_col = 'V'
    end_col = 'X'
    
    # Convert column letters to numbers (1-based)
    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col)
    
    # Find the last populated row in column N
    max_row = 1
    blankCount = 0
    for row in sheet["V"]:
        if row.value is not None:
            if row.row > max_row:
                max_row = row.row
        else:
            blankCount +=1
            if blankCount >= 10:
                break

    blankCount = 0
    for row in sheet["W"]:
        if row.value is not None:
            if row.row > max_row:
                max_row = row.row
        else:
            blankCount +=1

            if blankCount >= 10:
                break

    blankCount = 0
    for row in sheet["X"]:
        if row.value is not None:
            if row.row > max_row:
                max_row = row.row
        else:
            blankCount +=1

            if blankCount >= 10:
                break
    
    # Create a list to store our data
    data = []
    
    # Read the data from the sheet
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(start_col_idx, end_col_idx + 1):
            cell_value = sheet[f'{get_column_letter(col)}{row}'].value
            row_data.append(cell_value)
        data.append(row_data)
    

    # Get the header row (first row)
    headers = data[10]
    
    # Create DataFrame with the data, excluding the header row
    df = pd.DataFrame(data[11:], columns=headers)
    
    # Close the workbook
    wb.close()
    
    return df



